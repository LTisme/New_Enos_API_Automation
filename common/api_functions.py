"""
#-*- coding: utf-8 -*-

Created on 2021/1/14 12:32

@author: LT
"""
import os
import openpyxl
import xlwt
import pprint
import json
import re
import requests
import time     # 用来生成相应年月日名字 文件夹的
from requests_toolbelt import MultipartEncoder


def extract_data_into_a_json():
    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    wb = openpyxl.load_workbook('data.xlsx')
    sht = wb['Sheet1']

    data_list = []  # 字典列表
    info_list = []  # 信息列表，后续还得根据情况是否需要新创建

    STATION = ""  # 用来暂存空站名前最新的站名
    STATION_ID = ""  # 用来存最新的站ID
    ADDRESS = ""  # 用来存储最新的地址
    EXCEL_ID = ""  # 用来存储最新的,后续拓扑结构需要用到的每个站点对应的excel_id
    PORT = ""  # 用来存最新的端口号
    TEMP = ""  # 用来判断是否是同一个场站，它有一个暂存量的作用
    for row in range(2, sht.max_row + 1):  # 要包括进excel表里的最后一行
        """从里到外层层脱出"""
        if sht['C' + str(row)].value is not None and row >= 2 and sht['A' + str(row)].value is not None and \
                sht['J' + str(row)].value is not None:  # 如果设备名称不为空，站名不为空,端口号不为空
            STATION = sht['A' + str(row)].value  # 更新STATION
            STATION_ID = sht['B' + str(row)].value  # 更新STATION_ID
            ADDRESS = sht['K' + str(row)].value  # 更新地址
            EXCEL_ID = sht['L' + str(row)].value  # 更新excel_id
            PORT = sht['J' + str(row)].value  # 更新PORT
            device_name = sht['C' + str(row)].value  # 设备名
            modbus = sht['D' + str(row)].value  # 公共地址
            capacity = sht['E' + str(row)].value  # 容量
            rated_current = sht['F' + str(row)].value  # 额定电流
            CT = sht['G' + str(row)].value  # CT变比
            PT = sht['H' + str(row)].value  # PT变比
            manufacturer = sht['I' + str(row)].value  # 厂家名字
            port_num = sht['J' + str(row)].value  # 端口号
            # >>>>>>>>>>>>>>>>>>>>>>>基础信息分隔符
            site_ID = sht['B' + str(row)].value  # siteID
            station_name = sht['A' + str(row)].value  # 站点名字
            # >>>>>>>>>>>>>>>>>>>>>>>
            # 根据新信息创建一个新字典
            info_list_elem = {'device_name': device_name, 'modbus': modbus, 'capacity': capacity,
                              'rated_current': rated_current, 'CT': CT, 'PT': PT, 'manufacturer': manufacturer,
                              'port_num': port_num}  # 字典列表的字典元素中的信息列表中的元素
            if STATION != TEMP and TEMP != "":
                info_list = []
                info_list.append(info_list_elem)
            else:
                info_list.append(info_list_elem)
        elif sht['A' + str(row)].value is None and row <= sht.max_row:  # 若站名为空，且不大于最大行数
            if sht['J' + str(row)].value is not None:  # 如果站名为空，但端口号不为空
                PORT = sht['J' + str(row)].value  # 更新PORT
                device_name = sht['C' + str(row)].value  # 设备名
                modbus = sht['D' + str(row)].value  # 公共地址
                capacity = sht['E' + str(row)].value  # 容量
                rated_current = sht['F' + str(row)].value  # 额定电流
                CT = sht['G' + str(row)].value  # CT变比
                PT = sht['H' + str(row)].value  # PT变比
                manufacturer = sht['I' + str(row)].value  # 厂家名字
                port_num = sht['J' + str(row)].value  # 端口号
                # >>>>>>>>>>>>>>>>>>>>>>>基础信息分隔符
                site_ID = STATION_ID  # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<则用STATION_ID替代siteID
                station_name = STATION  # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<则用STATION替代站点名字
                # >>>>>>>>>>>>>>>>>>>>>>>
                # 根据新信息创建一个新字典
                info_list_elem = {'device_name': device_name, 'modbus': modbus, 'capacity': capacity,
                                  'rated_current': rated_current, 'CT': CT, 'PT': PT, 'manufacturer': manufacturer,
                                  'port_num': port_num}  # 字典列表的字典元素中的信息列表中的元素
                if STATION != TEMP and TEMP != "":
                    info_list = []
                    info_list.append(info_list_elem)
                else:
                    info_list.append(info_list_elem)
            elif sht['J' + str(row)].value is None:  # 如果站名为空，端口号也为空
                device_name = sht['C' + str(row)].value  # 设备名
                modbus = sht['D' + str(row)].value  # 公共地址
                capacity = sht['E' + str(row)].value  # 容量
                rated_current = sht['F' + str(row)].value  # 额定电流
                CT = sht['G' + str(row)].value  # CT变比
                PT = sht['H' + str(row)].value  # PT变比
                manufacturer = sht['I' + str(row)].value  # 厂家名字
                port_num = PORT  # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<则用PORT代替端口号
                # >>>>>>>>>>>>>>>>>>>>>>>基础信息分隔符
                site_ID = STATION_ID  # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<则用STATION_ID替代siteID
                station_name = STATION  # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<则用STATION替代站点名字
                # >>>>>>>>>>>>>>>>>>>>>>>
                # 根据新信息创建一个新字典
                info_list_elem = {'device_name': device_name, 'modbus': modbus, 'capacity': capacity,
                                  'rated_current': rated_current, 'CT': CT, 'PT': PT, 'manufacturer': manufacturer,
                                  'port_num': port_num}  # 字典列表的字典元素中的信息列表中的元素
                if STATION != TEMP and TEMP != "":
                    info_list = []
                    info_list.append(info_list_elem)
                else:
                    info_list.append(info_list_elem)
        else:
            print("表头下无数据")

        # 层层脱出==============================================层层脱出
        if TEMP != STATION:
            TEMP = STATION
            data_list_elem = {"站名": STATION, "siteID": STATION_ID, "address": ADDRESS, "excelid": EXCEL_ID,
                              "信息": info_list}  # 字典列表的字典元素
            data_list.append(data_list_elem)

    pprint.pprint(data_list)

    # 把结果保存为json文件

    with open('data.json', 'w', encoding='UTF-8') as file_obj:
        json.dump(data_list, file_obj, ensure_ascii=False)  # ensure_ascii=False 用来存储为真正的中文


def judgement(port_num):
    """
    用来判断端口号用的ip地址和sn号
    :param port_num: 传入的端口号
    :return: IP,SN 都是str型的
    """
    port_num = int(port_num)  # 先将传入的参数进行字符化处理
    if (30002 <= port_num <= 30200) or (30402 <= port_num <= 30500) \
            or (30501 <= port_num <= 30600) or (30801 <= port_num <= 31200):
        ip = '10.65.26.143'
        sn = 'de221685-5460-4c50-80ff-3d322eb5b019'
    elif (30201 <= port_num <= 30400) or (30601 <= port_num <= 30700) \
            or (31201 <= port_num <= 32000):
        ip = '10.65.26.144'
        sn = '90829856-5781-41fe-a6d4-d1ec5f16e548'
    elif 32001 <= port_num <= 32400:
        ip = '192.168.9.74'
        sn = '2462f05b-c90a-4a92-9469-43444d0297dd'
    elif 32401 <= port_num <= 32800:
        ip = '192.168.9.83'
        sn = 'a742eaa1-7348-472b-b32d-a9f00abdbda0'
    elif 32801 <= port_num <= 33200:
        ip = '192.168.9.86'
        sn = 'da5526ba-afc0-4eae-a7a5-fcd947126c5f'
    elif 33201 <= port_num <= 33600:
        ip = '192.168.9.88'
        sn = '584a2b58-8878-4c37-9144-cd4f427c7d04'
    elif 33601 <= port_num <= 34000:
        ip = '192.168.9.89'
        sn = '5cf81848-1122-4d0c-b665-2b44f1899a13'
    elif 34001 <= port_num <= 34400:
        ip = '192.168.9.94'
        sn = 'd3af993f-3f82-45ae-85e9-d8a093db4aff'
    elif 34401 <= port_num <= 34800:
        ip = '192.168.9.95'
        sn = 'da52699e-311f-4317-a675-12d07d59151d'
    elif 34801 <= port_num <= 35200:
        ip = '192.168.9.98'
        sn = '6e562d4f-9cfd-49b4-8c30-58b5d673b97d'
    elif 35201 <= port_num <= 35600:
        ip = '192.168.9.131'
        sn = '5cc7755c-a417-4cc7-9571-172753739f0d'
    elif 35601 <= port_num <= 36000:
        ip = '192.168.9.132'
        sn = 'b2c4ce75-4d77-445f-858d-650d24b53d44'
    elif 36001 <= port_num <= 36400:
        ip = '192.168.9.133'
        sn = '53d7c2e1-066b-4a63-aff9-48234d8d423b'
    elif 36401 <= port_num <= 36800:
        ip = '192.168.9.134'
        sn = 'c349eb79-e992-4a86-bb9b-f0983b13be7f'
    elif 36801 <= port_num <= 37200:
        ip = '192.168.9.135'
        sn = 'cc7d1a27-f180-477d-a361-f8d21988633e'
    elif 37201 <= port_num <= 37600:
        ip = '192.168.9.136'
        sn = 'db88f85c-d1cf-4a20-8521-3f62e4bcc335'
    elif 37601 <= port_num <= 38000:
        ip = '192.168.9.137'
        sn = '26eb74ff-dbf1-4bcc-ab21-b9f97af7c76a'
    elif 38001 <= port_num <= 38400:
        ip = '192.168.9.138'
        sn = '36c3de54-8547-41f1-8050-357fd9e6c972'
    elif 38401 <= port_num <= 38800:
        ip = '192.168.9.139'
        sn = '5ae573b7-5919-4337-a893-c7e0ab02b3c1'
    elif 38801 <= port_num <= 39200:
        ip = '192.168.9.140'
        sn = '3727f1f3-98a3-4e6e-b26e-c32936bcd7b2'
    elif 39201 <= port_num <= 39600:
        ip = '192.168.9.141'
        sn = '903dda4d-f78a-4780-97e3-00a0ac22dfdb'
    else:
        return False
    return ip, sn


def which_link(port_num, port_list, start_num=0):
    """
    判断该用哪个连接名字来进行下一步的操作
    :param port_num: 传入的端口号，用来查重
    :param port_list: 最开始应该为一个空列表，随着端口号的输入，而变得不空
    :param start_num: 起始后缀，默认为0；如果已有连接则需先去获得有多少连接数，再将此数传入进来
    :return: str
    """
    if port_num not in port_list:
        port_list.append(port_num)
        X = len(port_list) + start_num
    else:
        X = len(port_list) + start_num
    return "104转发%d" % X


def which_template(templates, manufacturer, CT, PT):
    """
    用输入的参数结合正则判定式来匹配到目标模板，当然要考虑到匹配不成功的情况该如何处理
    :param templates: 输入id与模板名对应的元组列表
    :param manufacturer: 厂家信息，不可或缺
    :param CT: 电流互感器变比
    :param PT: 电压互感器变比
    :return: str
    """
    if '南德电气' in manufacturer and PT != '1':
        # 南德电气20kv_PT200/1_CT600/5
        REGEX = re.compile(r'^(.*)(%s)(.*)(%s)(.*)(%s)$' % ('南德电气', PT, CT))  # 南德电气用的正则
        print('1')
    elif ('深圳中电' in manufacturer or '创力' in manufacturer or '南德电气' in manufacturer) and PT == '1' and CT != '1':
        # 创力目前没有带PT变比，深圳中电与创力用的同一个模板
        # 创力104转发YC78_1000/5
        REGEX = re.compile(r'^(.*)(%s)(.*)(_%s)$' % ('创力', CT))  # 带下划线可以区分2500/5与500/5的模糊
        print('2')
    elif '深圳中电' in manufacturer and CT == '1':
        # 深圳中电CT为1的，就得选温州电管家104转发YC78
        REGEX = re.compile(r'^温州电管家104转发YC78$')
        print('3')
    elif '佳和' in manufacturer and PT != '1':  # 佳和带PT的变比
        # 温州佳和104转发YC91_100/5_PT100/1
        REGEX = re.compile(r'^(.*)(%s)(.*)(_%s)(.*)(_PT%s)$' % ('佳和', CT, PT))
        print('4')
    elif ('佳和' in manufacturer and '水电' not in manufacturer) and PT == '1':  # 佳和不带PT的变比，最好带一个非贪心匹配，这样可以不带入水电后缀
        # 温州佳和104转发YC91_100/5
        REGEX = re.compile(r'^(.*)(%s)(.*)(_%s)$' % ('佳和', CT))
        print('5')
    elif '水电' in manufacturer and PT == '1':  # 水电目前只有佳和水电才有
        # 温州佳和104转发YC91_500/5_水电
        REGEX = re.compile(r'^(.*)(%s)(.*)(_%s)(.*)(_水电)$' % ('佳和', CT))
        print('6')
    else:
        print(f'匹配失败')
        return 'failed'  # 正则式匹配失败，需要外界判断
    final = ""
    for target in templates:
        if REGEX.match(target[1]):
            final = str(target[0])  # 若成功则返回对应的id，也就是F12option中的value值
            break
    if final != "":
        print(f'匹配成功')
        return final
    else:
        print(f'匹配失败')
        return 'failed'


def synchronize_topo(info_dict):
    """
    将传入参数写入拓扑结构模板并上传同步
    :param info_dict: 含有必要信息的字典
    :return: str 创建的excel表的绝对路径名
    """
    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    time_now = time.strftime("%Y-%m-%d", time.localtime())      # 获得当前的年月日
    if not os.path.exists(r'synchronization_excel_files/{}'.format(time_now)):
        os.makedirs(r'synchronization_excel_files/{}'.format(time_now))
    save_path = r'synchronization_excel_files/{}/{}.xls'.format(time_now, info_dict['station_name'])

    def judge(name):
        """
        用来判断设备是否需要纳入首页电量计算的函数
        :param name: 设备名字
        :return: bool
        """
        possible_list = ['总开关', '总电流', '总进线', '总路', '进线', '发电', '红外', '水电', '主回路', '主开关']   # 这个列表暂时只能想到这么多
        for elem in possible_list:
            if elem in name and '出线' not in name:   # 避免如 低压进线出线回路1 这样的设备名导致的误解
                return True
        return False

    def post_excel(name, excel_id, real_path):  # 用来上传刚保存的excel文件到目标url
        """
        内置上传excel文件函数
        :param name: 站点名
        :param excel_id: 对应的excel_id，虽然它这要求上传的参数是siteId
        :param real_path: 文件的绝对路径
        :return: 返回response的json
        """
        m = MultipartEncoder(
            fields={"siteId": excel_id,
                    'file': (f'{name}.xls', open(real_path, 'rb'), 'application/vnd.ms-excel')},
            boundary='------WebKitFormBoundarytixmuvUD9Ef6p61F'
        )
        res = requests.request("POST", 'http://122.228.156.194:8081/backend/company/upload/elementImport', data=m,
                               headers={'Content-Type': m.content_type}, )
        return res.json()

    # xlwt部分
    newbook = xlwt.Workbook(encoding="utf-8", style_compression=0)  # 创建一个工作簿对象
    # 在工作簿中创建4个工作表对象
    sheet1 = newbook.add_sheet('开关柜', cell_overwrite_ok=True)  # 设置参数为可覆盖
    sheet2 = newbook.add_sheet('配电房', cell_overwrite_ok=True)  # 设置参数为可覆盖
    sheet3 = newbook.add_sheet('回路', cell_overwrite_ok=True)  # 设置参数为可覆盖
    sheet4 = newbook.add_sheet('谐波表', cell_overwrite_ok=True)  # 设置参数为可覆盖

    # 开关柜表的表头
    tag_1 = ['名称', '额定电压', '父设备类型', '父设备', 'mdmid', '开关柜编号', '开关柜型号']
    # 只需要在content_1[3]中填入站点名字
    content_1 = ['1AA1', '380', '配电房', '', '0', '', '']
    # 配电房表的表头
    tag_2 = ['名称', '额定电压', '父设备类型', '父设备', 'mdmid', '站点名', '配电房地址', '现场网络情况', '信号强度', '投运日期']
    # 在content_2[0]、content_2[5]中填入站点名字，在content_2[6]中填入地址
    content_2 = ['', '380', '-', 'No', '0', '', '', '无线', '无线50dB', '']
    # 回路表的表头
    tag_3 = ['名称', '额定电压', '父设备类型', '父设备', 'mdmid', '负荷用途', '标签(建筑类)', '标签(工业类)']
    # 只需要在content_3[0]中填入设备名字——这个可有多条，条数用len(dict['devices_list'])来判断
    content_3 = ['', '380', '开关柜', '1AA1', '0', '', '', '']
    # 谐波表的表头
    tag_4 = ['名称', '额定电压', '父设备类型', '父设备', 'mdmid', '电能质量监测点-功率因数', '电压偏差', '三相电流不平衡', '电压偏差-标称类别', '电压偏差-标称值（V）',
             '是否纳入首页电量统计', '谐波电压', '谐波电流']
    # 在content_4[0]、content_4[3]中填入设备名，在content_4[4]中填入设备对应的objectid，在在content_4[10]中用正则判断一下是否纳入首页电量计算——也是多条
    content_4 = ['', '380', '回路', '', '', '是', '是', '是', '线电压', '380', '', '是', '是']

    # 先统一将4张表的表头写好
    for i in range(0, len(tag_1)):
        sheet1.write(0, i, tag_1[i])
    for i in range(0, len(tag_2)):
        sheet2.write(0, i, tag_2[i])
    for i in range(0, len(tag_3)):
        sheet3.write(0, i, tag_3[i])
    for i in range(0, len(tag_4)):
        sheet4.write(0, i, tag_4[i])
    # 再去统一写下面变动的内容，这个时候需要解析传参了
    start_num = 1   # 因为第一行都被表头占据了，所以要从1而不是0开始了

    content_1[3] = info_dict['station_name']
    for i in range(0, len(content_1)):      # 写 开关柜 表
        sheet1.write(start_num, i, content_1[i])

    content_2[0], content_2[5], content_2[6] = info_dict['station_name'], info_dict['station_name'], info_dict['address']
    for i in range(0, len(content_2)):      # 写 配电房 表
        sheet2.write(start_num, i, content_2[i])

    # 接下来的内容行数是不固定的
    nums = len(info_dict['devices_list'])
    for i in range(0, nums):    # 写 回路 表
        content_3[0] = info_dict['devices_list'][i]['name']
        for j in range(0, len(content_3)):
            sheet3.write(start_num + i, j, content_3[j])
    for i in range(0, nums):    # 写 谐波表 表
        if judge(info_dict['devices_list'][i]['name']):
            content_4[10] = '是'
        else:
            content_4[10] = '否'
        content_4[0], content_4[3], content_4[4] = info_dict['devices_list'][i]['name'], info_dict['devices_list'][i]['name'], info_dict['devices_list'][i]['objectID']
        for j in range(0, len(content_4)):
            sheet4.write(start_num + i, j, content_4[j])

    newbook.save(save_path)     # 保存修改的工作表
    abs_path = os.path.realpath(save_path)
    return post_excel(info_dict['station_name'], info_dict['excel_id'], abs_path)   # 返回response的结果


if __name__ == '__main__':
    # test = {
    #     "retCode": 10000,
    #     "errMsg": "",
    #     "data": [
    #         {
    #             "id": 4109,
    #             "deviceName": "温州小微园区23YC",
    #             "brand": "极熵",
    #             "model": "jishang_23",
    #             "innerVer": "v1.0",
    #             "driverId": 11777,
    #             "protocolId": 1257,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14046",
    #                 "14047"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1565684625000,
    #             "updatedDate": 1584521828000
    #         },
    #         {
    #             "id": 4112,
    #             "deviceName": "温州小微园区24YC",
    #             "brand": "极熵",
    #             "model": "jishang_24",
    #             "innerVer": "v1.0",
    #             "driverId": 11776,
    #             "protocolId": 1257,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14046",
    #                 "14048"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1567735360000,
    #             "updatedDate": 1584518069000
    #         },
    #         {
    #             "id": 4114,
    #             "deviceName": "温州电管家104转发YC78",
    #             "brand": "电管家",
    #             "model": "温州电管家YC78",
    #             "innerVer": "v1.0",
    #             "driverId": 11911,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14238"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1568603158000,
    #             "updatedDate": 1603433601000
    #         },
    #         {
    #             "id": 4123,
    #             "deviceName": "温州佳和104转发YC91",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91",
    #             "innerVer": "v1.0",
    #             "driverId": 11726,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14078"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1572833573000,
    #             "updatedDate": 1576050330000
    #         },
    #         {
    #             "id": 4130,
    #             "deviceName": "温州墨熵104转发YC110",
    #             "brand": "墨熵",
    #             "model": "温州墨熵YC110",
    #             "innerVer": "v1.0",
    #             "driverId": 11709,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14079",
    #                 "14094"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1573094044000,
    #             "updatedDate": 1573203075000
    #         },
    #         {
    #             "id": 4135,
    #             "deviceName": "温州佳和104转发YC91_800/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_800/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11818,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14167",
    #                 "14159"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1574905445000,
    #             "updatedDate": 1592915255000
    #         },
    #         {
    #             "id": 4136,
    #             "deviceName": "温州佳和104转发YC91_400/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_400/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11804,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14106",
    #                 "14156"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1574905586000,
    #             "updatedDate": 1592021731000
    #         },
    #         {
    #             "id": 4137,
    #             "deviceName": "温州佳和104转发YC91_500/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11808,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14105",
    #                 "14160"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1574905771000,
    #             "updatedDate": 1592022110000
    #         },
    #         {
    #             "id": 4138,
    #             "deviceName": "温州佳和104转发YC91_600/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11820,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14147"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576130942000,
    #             "updatedDate": 1592920411000
    #         },
    #         {
    #             "id": 4139,
    #             "deviceName": "温州佳和104转发YC91_100/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_100/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11794,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14146"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576131245000,
    #             "updatedDate": 1592020816000
    #         },
    #         {
    #             "id": 4140,
    #             "deviceName": "温州佳和104转发YC91_1000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11816,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14165",
    #                 "14142"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576132724000,
    #             "updatedDate": 1592725492000
    #         },
    #         {
    #             "id": 4141,
    #             "deviceName": "温州佳和104转发YC91_75/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_75/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11823,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14170"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576132783000,
    #             "updatedDate": 1595035248000
    #         },
    #         {
    #             "id": 4142,
    #             "deviceName": "温州佳和104转发YC91_50/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_50/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11806,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14158"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576132840000,
    #             "updatedDate": 1592021880000
    #         },
    #         {
    #             "id": 4143,
    #             "deviceName": "温州佳和104转发YC91_150/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_150/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11792,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14144"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576132883000,
    #             "updatedDate": 1592020371000
    #         },
    #         {
    #             "id": 4144,
    #             "deviceName": "温州佳和104转发YC91_200/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11793,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14145"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576132936000,
    #             "updatedDate": 1592020433000
    #         },
    #         {
    #             "id": 4149,
    #             "deviceName": "温州佳和104转发YC91_2000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_2000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11803,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14155"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576931156000,
    #             "updatedDate": 1592021589000
    #         },
    #         {
    #             "id": 4150,
    #             "deviceName": "温州佳和104转发YC91_300/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_300/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11805,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14157"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576931380000,
    #             "updatedDate": 1592021798000
    #         },
    #         {
    #             "id": 4151,
    #             "deviceName": "温州佳和104转发YC91_1500/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11802,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14154"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1576993839000,
    #             "updatedDate": 1592021515000
    #         },
    #         {
    #             "id": 4152,
    #             "deviceName": "温州佳和104转发YC91_1200/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11801,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14153"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1577023522000,
    #             "updatedDate": 1592021405000
    #         },
    #         {
    #             "id": 4153,
    #             "deviceName": "温州佳和104转发YC91_250/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_250/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11800,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14152"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1577098246000,
    #             "updatedDate": 1592021323000
    #         },
    #         {
    #             "id": 4154,
    #             "deviceName": "温州佳和104转发YC91_2500/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_2500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11799,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14151"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1577115261000,
    #             "updatedDate": 1592021219000
    #         },
    #         {
    #             "id": 4157,
    #             "deviceName": "温州佳和104转发YC91_1600/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11798,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14150"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1577177975000,
    #             "updatedDate": 1592021150000
    #         },
    #         {
    #             "id": 4158,
    #             "deviceName": "modbus-rtu测试",
    #             "brand": "测试",
    #             "model": "测试",
    #             "innerVer": "v1.0",
    #             "driverId": 11890,
    #             "protocolId": 1270,
    #             "driverName": "",
    #             "driverProto": "ModbusRTU",
    #             "confPathList": [
    #                 "14125",
    #                 "14126"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1577271625000,
    #             "updatedDate": 1599549346000
    #         },
    #         {
    #             "id": 4160,
    #             "deviceName": "50KW光伏逆变器",
    #             "brand": "光伏逆变器",
    #             "model": "50KW光伏逆变器",
    #             "innerVer": "v1.0",
    #             "driverId": 11772,
    #             "protocolId": 1271,
    #             "driverName": "",
    #             "driverProto": "ModbusRTU",
    #             "confPathList": [
    #                 "14129",
    #                 "14132"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 234,
    #             "domainId": 418,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1577277216000,
    #             "updatedDate": 1578020001000
    #         },
    #         {
    #             "id": 4161,
    #             "deviceName": "温州佳和104转发YC91_3000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_3000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11797,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14149"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1579401972000,
    #             "updatedDate": 1592021071000
    #         },
    #         {
    #             "id": 4163,
    #             "deviceName": "温州电管家104转发YC78_copy",
    #             "brand": "电管家",
    #             "model": "温州电管家YC78_1587778048857",
    #             "innerVer": "v1.0",
    #             "driverId": 11780,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14134",
    #                 "14072"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1587778048000,
    #             "updatedDate": 1587778255000
    #         },
    #         {
    #             "id": 4164,
    #             "deviceName": "温州佳和104转发YC91_70/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_70/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11796,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14148"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1588227488000,
    #             "updatedDate": 1592020998000
    #         },
    #         {
    #             "id": 4165,
    #             "deviceName": "温州电管家104转发YC79/78",
    #             "brand": "电管家",
    #             "model": "温州电管家YC78_1591841500785",
    #             "innerVer": "v1.0",
    #             "driverId": 11783,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14100",
    #                 "14072"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1591841500000,
    #             "updatedDate": 1591841500000
    #         },
    #         {
    #             "id": 4166,
    #             "deviceName": "温州佳和104转发YC91_5000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_5000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11810,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14105",
    #                 "14161"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1592099702000,
    #             "updatedDate": 1592099804000
    #         },
    #         {
    #             "id": 4167,
    #             "deviceName": "温州佳和104转发YC91_750/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_750/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11813,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14105",
    #                 "14163"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1592188553000,
    #             "updatedDate": 1592202116000
    #         },
    #         {
    #             "id": 4168,
    #             "deviceName": "温州佳和104转发YC91_4000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_4000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11815,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14106",
    #                 "14164"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1592208730000,
    #             "updatedDate": 1592208779000
    #         },
    #         {
    #             "id": 4169,
    #             "deviceName": "创力104转发YC78_800/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_800/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11881,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14217"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1594608096000,
    #             "updatedDate": 1598258360000
    #         },
    #         {
    #             "id": 4170,
    #             "deviceName": "创力104转发YC78_4000/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_4000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11875,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14211"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595381811000,
    #             "updatedDate": 1598257492000
    #         },
    #         {
    #             "id": 4171,
    #             "deviceName": "创力104转发YC78_2000/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_2000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11873,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14208"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595382553000,
    #             "updatedDate": 1598254466000
    #         },
    #         {
    #             "id": 4172,
    #             "deviceName": "创力104转发YC78_3000/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_3000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11874,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14210"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595382623000,
    #             "updatedDate": 1598257287000
    #         },
    #         {
    #             "id": 4173,
    #             "deviceName": "创力104转发YC78_400/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_400/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11880,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14216"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595465103000,
    #             "updatedDate": 1598258291000
    #         },
    #         {
    #             "id": 4174,
    #             "deviceName": "创力104转发YC78_1500/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11879,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14215"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595471387000,
    #             "updatedDate": 1598258187000
    #         },
    #         {
    #             "id": 4175,
    #             "deviceName": "创力104转发YC78_600/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11878,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14214"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1595578550000,
    #             "updatedDate": 1598258119000
    #         },
    #         {
    #             "id": 4176,
    #             "deviceName": "创力104转发YC78_1000/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11877,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14213"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1596004531000,
    #             "updatedDate": 1598258017000
    #         },
    #         {
    #             "id": 4177,
    #             "deviceName": "创力104转发YC78_300/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_300/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11900,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14229"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1596177322000,
    #             "updatedDate": 1601169637000
    #         },
    #         {
    #             "id": 4178,
    #             "deviceName": "创力104转发YC78_500/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11872,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14207"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1596245128000,
    #             "updatedDate": 1598251483000
    #         },
    #         {
    #             "id": 4180,
    #             "deviceName": "创力104转发YC78_75/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_75/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11869,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14204"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1596943833000,
    #             "updatedDate": 1598238881000
    #         },
    #         {
    #             "id": 4181,
    #             "deviceName": "创力104转发YC78_200/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11871,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14206"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597111882000,
    #             "updatedDate": 1598251325000
    #         },
    #         {
    #             "id": 4182,
    #             "deviceName": "创力104转发YC78_5000/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_5000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11870,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14205"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597115190000,
    #             "updatedDate": 1598251237000
    #         },
    #         {
    #             "id": 4183,
    #             "deviceName": "温州佳和104转发YC91_6000/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_6000/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11860,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14195"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597391095000,
    #             "updatedDate": 1597391151000
    #         },
    #         {
    #             "id": 4184,
    #             "deviceName": "创力104转发YC78_750/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_750/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11868,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14203"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597805086000,
    #             "updatedDate": 1598238558000
    #         },
    #         {
    #             "id": 4185,
    #             "deviceName": "创力104转发YC78_1200/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11867,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14202"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597809054000,
    #             "updatedDate": 1598238414000
    #         },
    #         {
    #             "id": 4186,
    #             "deviceName": "创力104转发YC78_1200/5测试",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1200/5测试",
    #             "innerVer": "v1.0",
    #             "driverId": 11866,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14201"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1597996993000,
    #             "updatedDate": 1597997811000
    #         },
    #         {
    #             "id": 4187,
    #             "deviceName": "极熵进线表模板",
    #             "brand": "极熵",
    #             "model": "MEA220",
    #             "innerVer": "V1.0",
    #             "driverId": 11977,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14219",
    #                 "14279"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "JSCS",
    #             "createdDate": 1598328254000,
    #             "updatedDate": 1606986949000
    #         },
    #         {
    #             "id": 4188,
    #             "deviceName": "创力104转发YC78_250/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_250/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11885,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14220"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1599020141000,
    #             "updatedDate": 1599020223000
    #         },
    #         {
    #             "id": 4189,
    #             "deviceName": "创力104转发YC78_2500/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_2500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11887,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14221"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1599020316000,
    #             "updatedDate": 1599020424000
    #         },
    #         {
    #             "id": 4190,
    #             "deviceName": "创力104转发YC78_3200/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_3200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11889,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14222"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1599447919000,
    #             "updatedDate": 1600046902000
    #         },
    #         {
    #             "id": 4191,
    #             "deviceName": "创力104转发YC78_100/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_100/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11892,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14223"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1599735474000,
    #             "updatedDate": 1600047087000
    #         },
    #         {
    #             "id": 4192,
    #             "deviceName": "创力104转发YC78_150/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_150/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11894,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14224"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1599735947000,
    #             "updatedDate": 1600047238000
    #         },
    #         {
    #             "id": 4194,
    #             "deviceName": "温州佳和104转发YC91_30/5",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_30/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11897,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14225"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1600331527000,
    #             "updatedDate": 1600331603000
    #         },
    #         {
    #             "id": 4195,
    #             "deviceName": "创力104转发YC78_1250/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1250/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11899,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14228"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1600658762000,
    #             "updatedDate": 1600658903000
    #         },
    #         {
    #             "id": 4197,
    #             "deviceName": "温州佳和104转发YC91_600/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_600/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11905,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14231"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1601172716000,
    #             "updatedDate": 1601173138000
    #         },
    #         {
    #             "id": 4198,
    #             "deviceName": "温州佳和104转发YC91_1500/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1500/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11904,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14230"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1601172839000,
    #             "updatedDate": 1601173075000
    #         },
    #         {
    #             "id": 4199,
    #             "deviceName": "温州佳和104转发YC91_1000/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_1000/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11908,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14232"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1601184821000,
    #             "updatedDate": 1601184942000
    #         },
    #         {
    #             "id": 4200,
    #             "deviceName": "温州佳和104转发YC91_800/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_800/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11909,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14233"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1601184839000,
    #             "updatedDate": 1601184996000
    #         },
    #         {
    #             "id": 4201,
    #             "deviceName": "红外表头YC5_600/5",
    #             "brand": "深圳中电",
    #             "model": "红外表头YC5_600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11910,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14235",
    #                 "14234"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1602750110000,
    #             "updatedDate": 1602750535000
    #         },
    #         {
    #             "id": 4202,
    #             "deviceName": "创力104转发YC78_50/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_50/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11913,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14239"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1604381793000,
    #             "updatedDate": 1604381910000
    #         },
    #         {
    #             "id": 4203,
    #             "deviceName": "温州佳和104转发YC91_2000/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_2000/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11915,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14241"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1604547378000,
    #             "updatedDate": 1604547624000
    #         },
    #         {
    #             "id": 4204,
    #             "deviceName": "创力104转发YC78_30/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_30/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11917,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14242"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1604644133000,
    #             "updatedDate": 1604644189000
    #         },
    #         {
    #             "id": 4205,
    #             "deviceName": "南德电气20kv_PT200/1_CT50/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT50/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11923,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14247"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1604646022000,
    #             "updatedDate": 1605674474000
    #         },
    #         {
    #             "id": 4206,
    #             "deviceName": "南德电气20kv_PT200/1_CT30/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT30/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11922,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14246"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1604646938000,
    #             "updatedDate": 1605674453000
    #         },
    #         {
    #             "id": 4207,
    #             "deviceName": "南德电气20kv_PT200/1_CT75/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT75/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11925,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14248"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605599754000,
    #             "updatedDate": 1605674503000
    #         },
    #         {
    #             "id": 4208,
    #             "deviceName": "南德电气20kv_PT200/1_CT20/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT20/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11927,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14249"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605600175000,
    #             "updatedDate": 1605674381000
    #         },
    #         {
    #             "id": 4209,
    #             "deviceName": "南德电气20kv_PT200/1_CT25/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT25/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11929,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14250"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605600362000,
    #             "updatedDate": 1605674415000
    #         },
    #         {
    #             "id": 4210,
    #             "deviceName": "南德电气20kv_PT200/1_CT100/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT100/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11955,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14266"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605600446000,
    #             "updatedDate": 1605837568000
    #         },
    #         {
    #             "id": 4211,
    #             "deviceName": "南德电气20kv_PT200/1_CT150/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT150/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11934,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14254"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605660151000,
    #             "updatedDate": 1605674595000
    #         },
    #         {
    #             "id": 4212,
    #             "deviceName": "南德电气20kv_PT200/1_CT400/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT400/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11933,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14252"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605662634000,
    #             "updatedDate": 1605674556000
    #         },
    #         {
    #             "id": 4213,
    #             "deviceName": "南德电气20kv_PT200/1_CT600/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11936,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14255"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605662829000,
    #             "updatedDate": 1605674641000
    #         },
    #         {
    #             "id": 4214,
    #             "deviceName": "南德电气20kv_PT100/1_CT30/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT30/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11938,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14256"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663018000,
    #             "updatedDate": 1605674705000
    #         },
    #         {
    #             "id": 4215,
    #             "deviceName": "南德电气20kv_PT100/1_CT50/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT50/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11940,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14257"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663124000,
    #             "updatedDate": 1605674729000
    #         },
    #         {
    #             "id": 4216,
    #             "deviceName": "南德电气20kv_PT100/1_CT75/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT75/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11944,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14259"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663172000,
    #             "updatedDate": 1605674744000
    #         },
    #         {
    #             "id": 4217,
    #             "deviceName": "南德电气20kv_PT100/1_CT100/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT100/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11943,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14258"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663233000,
    #             "updatedDate": 1605674762000
    #         },
    #         {
    #             "id": 4218,
    #             "deviceName": "南德电气20kv_PT100/1_CT150/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT150/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11946,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14260"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663386000,
    #             "updatedDate": 1605674778000
    #         },
    #         {
    #             "id": 4219,
    #             "deviceName": "南德电气20kv_PT100/1_CT200/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT200/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11948,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14261"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663440000,
    #             "updatedDate": 1605674792000
    #         },
    #         {
    #             "id": 4220,
    #             "deviceName": "南德电气20kv_PT100/1_CT300/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT300/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11950,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14262"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663488000,
    #             "updatedDate": 1605674805000
    #         },
    #         {
    #             "id": 4221,
    #             "deviceName": "南德电气20kv_PT100/1_CT400/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT400/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11952,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14263"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605663529000,
    #             "updatedDate": 1605674818000
    #         },
    #         {
    #             "id": 4222,
    #             "deviceName": "南德电气20kv_PT200/1_CT40/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT200/1_CT40/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11954,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14264"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1605745970000,
    #             "updatedDate": 1605746069000
    #         },
    #         {
    #             "id": 4223,
    #             "deviceName": "温州佳和104转发YC91_400/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和YC91_400/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11957,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14267"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606272763000,
    #             "updatedDate": 1606272853000
    #         },
    #         {
    #             "id": 4224,
    #             "deviceName": "温州佳和104转发YC91_500/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_500/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11959,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14268"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606288354000,
    #             "updatedDate": 1606288387000
    #         },
    #         {
    #             "id": 4225,
    #             "deviceName": "温州佳和104转发YC91_200/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_200/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11961,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14269"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606786291000,
    #             "updatedDate": 1606786353000
    #         },
    #         {
    #             "id": 4226,
    #             "deviceName": "创力104转发YC78_1600/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_1600/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11963,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14270"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606788916000,
    #             "updatedDate": 1606788978000
    #         },
    #         {
    #             "id": 4227,
    #             "deviceName": "测试模板11111",
    #             "brand": "极熵",
    #             "model": "测试模板11111",
    #             "innerVer": "v1.0",
    #             "driverId": 11967,
    #             "protocolId": 1257,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14272",
    #                 "14273"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1606889081000,
    #             "updatedDate": 1606960701000
    #         },
    #         {
    #             "id": 4228,
    #             "deviceName": "温州电管家104转发YC78_1",
    #             "brand": "电管家",
    #             "model": "温州电管家YC78_1",
    #             "innerVer": "v1.0",
    #             "driverId": 11969,
    #             "protocolId": 1257,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14274",
    #                 "14238"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606962765000,
    #             "updatedDate": 1606962847000
    #         },
    #         {
    #             "id": 4229,
    #             "deviceName": "南德电气20kv_PT100/1_CT250/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT250/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11971,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14275"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606977266000,
    #             "updatedDate": 1606977331000
    #         },
    #         {
    #             "id": 4230,
    #             "deviceName": "南德电气20kv_PT100/1_CT500/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT500/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11973,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14276"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1606977544000,
    #             "updatedDate": 1606977591000
    #         },
    #         {
    #             "id": 4231,
    #             "deviceName": "变电站测温RTU规约测试",
    #             "brand": "测试",
    #             "model": "变电站测温RTU规约测试",
    #             "innerVer": "v1.0",
    #             "driverId": 11983,
    #             "protocolId": 1270,
    #             "driverName": "",
    #             "driverProto": "ModbusRTU",
    #             "confPathList": [
    #                 "14125",
    #                 "14286"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 377,
    #             "parentDomainId": 0,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1606986147000,
    #             "updatedDate": 1607586289000
    #         },
    #         {
    #             "id": 4232,
    #             "deviceName": "温州佳和104转发YC91_750/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_750/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11981,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14283"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1607414397000,
    #             "updatedDate": 1607414438000
    #         },
    #         {
    #             "id": 4233,
    #             "deviceName": "温州佳和104转发YC91_1200/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_1200/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11985,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14287"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1607997233000,
    #             "updatedDate": 1607997280000
    #         },
    #         {
    #             "id": 4234,
    #             "deviceName": "创力104转发YC78_450/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_450/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11987,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14288"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608086589000,
    #             "updatedDate": 1608086633000
    #         },
    #         {
    #             "id": 4235,
    #             "deviceName": "温州佳和104转发YC91_100/5_PT100/1",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_100/5_PT100/1",
    #             "innerVer": "v1.0",
    #             "driverId": 11990,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14289"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608351355000,
    #             "updatedDate": 1608351563000
    #         },
    #         {
    #             "id": 4236,
    #             "deviceName": "温州佳和104转发YC91_1000/5_PT100/1",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_1000/5_PT100/1",
    #             "innerVer": "v1.0",
    #             "driverId": 11991,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14290"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608351371000,
    #             "updatedDate": 1608351615000
    #         },
    #         {
    #             "id": 4237,
    #             "deviceName": "南德电气20kv_PT100/1_CT40/5",
    #             "brand": "创力",
    #             "model": "南德电气20kv_PT100/1_CT40/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11993,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14291"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608358135000,
    #             "updatedDate": 1608358173000
    #         },
    #         {
    #             "id": 4238,
    #             "deviceName": "创力104转发YC78_630/5",
    #             "brand": "创力",
    #             "model": "创力104转发YC78_630/5",
    #             "innerVer": "v1.0",
    #             "driverId": 11995,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14166",
    #                 "14292"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608513853000,
    #             "updatedDate": 1608513908000
    #         },
    #         {
    #             "id": 4239,
    #             "deviceName": "温州佳和104转发YC91_200/5_PT100/1",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_200/5_PT100/1",
    #             "innerVer": "v1.0",
    #             "driverId": 11997,
    #             "protocolId": 1308,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14168",
    #                 "14293"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "TSKJ",
    #             "createdDate": 1608794014000,
    #             "updatedDate": 1608794133000
    #         },
    #         {
    #             "id": 4240,
    #             "deviceName": "温州佳和104转发YC91_300/5_水电",
    #             "brand": "佳和",
    #             "model": "温州佳和104转发YC91_300/5_水电",
    #             "innerVer": "v1.0",
    #             "driverId": 11999,
    #             "protocolId": 1264,
    #             "driverName": "",
    #             "driverProto": "IEC104",
    #             "confPathList": [
    #                 "14104",
    #                 "14294"
    #             ],
    #             "categoryId": 29,
    #             "cimObjectTypeId": 227,
    #             "domainId": 378,
    #             "parentDomainId": 377,
    #             "namespace": "1859febe5ce70000",
    #             "owner": "shujun.wu",
    #             "createdDate": 1610415670000,
    #             "updatedDate": 1610415735000
    #         }
    #     ]
    # }
    # result = which_template(test, '南德电气', '400/5', '100/1')
    # print(result)
    # extract_data_into_a_json()
    example = {
        'station_name': '浙江帅帅电气科技',
        'address': '浙江省温州市平阳县万全镇工业园区',
        'site_id': '我是siteid',
        'excel_id': '我是excelid',
        'devices_list': [{'CT': '500/5',
                          'PT': '1',
                          'manufacturer': '南德电气',
                          'modbus': '1',
                          'name': '进线1',
                          'objectID': '25781d0897006000',
                          'portnum': '39600'},
                         {'CT': '400/5',
                          'PT': '100/1',
                          'manufacturer': '南德电气',
                          'modbus': '2',
                          'name': '即删设备1.1',
                          'objectID': '25781d0aa4802000',
                          'portnum': '39600'},
                         {'CT': '300/5',
                          'PT': '1',
                          'manufacturer': '创力',
                          'modbus': '3',
                          'name': '即删设备1.2',
                          'objectID': '25781d0c50002000',
                          'portnum': '39600'},
                         {'CT': '1500/5',
                          'PT': '1',
                          'manufacturer': '南德电气',
                          'modbus': '1',
                          'name': '即删设备2',
                          'objectID': '25781d0ea6002000',
                          'portnum': '39599'},
                         {'CT': '600/5',
                          'PT': '200/1',
                          'manufacturer': '南德电气',
                          'modbus': '1',
                          'name': '即删设备3',
                          'objectID': '25781d1084000000',
                          'portnum': '39598'},
                         {'CT': '1000/5',
                          'PT': '1',
                          'manufacturer': '南德电气',
                          'modbus': '1',
                          'name': '即删设备4',
                          'objectID': '25781d128f802000',
                          'portnum': '38888'}]
    }
    xc = synchronize_topo(example)
    print(xc)
